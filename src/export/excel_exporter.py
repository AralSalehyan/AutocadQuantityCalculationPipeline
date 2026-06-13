import json
from pathlib import Path

from src.geometry.primitives import DetectionObject, QuantityItem, to_plain


class ExcelExporter:
    def write(
        self,
        output_path: Path,
        quantities: list[QuantityItem],
        raw_detections: list[DetectionObject],
        merged_objects: list[DetectionObject],
        warnings: list[str],
    ) -> Path:
        try:
            return self._write_with_openpyxl(output_path, quantities, raw_detections, merged_objects, warnings)
        except ImportError:
            return self._write_minimal_xlsx(output_path, quantities, raw_detections, merged_objects, warnings)

    def _write_with_openpyxl(
        self,
        output_path: Path,
        quantities: list[QuantityItem],
        raw_detections: list[DetectionObject],
        merged_objects: list[DetectionObject],
        warnings: list[str],
    ) -> Path:
        from openpyxl import Workbook

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        _append_sheet(workbook.create_sheet("Room Areas"), _room_rows(quantities))
        _append_sheet(workbook.create_sheet("Openings"), _opening_rows(quantities))
        _append_sheet(workbook.create_sheet("Wall Quantities"), _wall_rows(quantities))
        _append_sheet(workbook.create_sheet("Raw Detections"), _object_rows(raw_detections))
        _append_sheet(workbook.create_sheet("Merged Objects"), _object_rows(merged_objects))
        _append_sheet(workbook.create_sheet("Warnings"), [["Warning"], *[[item] for item in warnings]])
        workbook.save(output_path)
        return output_path

    def _write_minimal_xlsx(
        self,
        output_path: Path,
        quantities: list[QuantityItem],
        raw_detections: list[DetectionObject],
        merged_objects: list[DetectionObject],
        warnings: list[str],
    ) -> Path:
        import html
        import zipfile

        output_path.parent.mkdir(parents=True, exist_ok=True)
        sheets = {
            "Room Areas": _room_rows(quantities),
            "Openings": _opening_rows(quantities),
            "Wall Quantities": _wall_rows(quantities),
            "Raw Detections": _object_rows(raw_detections),
            "Merged Objects": _object_rows(merged_objects),
            "Warnings": [["Warning"], *[[item] for item in warnings]],
        }

        def sheet_xml(rows: list[list[object]]) -> str:
            body = []
            for r_index, row in enumerate(rows, start=1):
                cells = []
                for c_index, value in enumerate(row, start=1):
                    ref = f"{chr(64 + c_index)}{r_index}"
                    cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{html.escape(str(value))}</t></is></c>')
                body.append(f'<row r="{r_index}">{"".join(cells)}</row>')
            return '<?xml version="1.0" encoding="UTF-8"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>' + "".join(body) + "</sheetData></worksheet>"

        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", _content_types(len(sheets)))
            archive.writestr("_rels/.rels", _rels())
            archive.writestr("xl/workbook.xml", _workbook_xml(list(sheets)))
            archive.writestr("xl/_rels/workbook.xml.rels", _workbook_rels(len(sheets)))
            for idx, rows in enumerate(sheets.values(), start=1):
                archive.writestr(f"xl/worksheets/sheet{idx}.xml", sheet_xml(rows))
        return output_path


def _append_sheet(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)


def _room_rows(quantities: list[QuantityItem]) -> list[list[object]]:
    header = ["Object ID", "Room Name", "Area", "Unit", "Perimeter", "Perimeter Unit", "Confidence", "Source"]
    rows = [header]
    area_items = [item for item in quantities if item.category == "room_area"]
    perimeter_items = [item for item in quantities if item.category == "room_perimeter"]
    perimeter_by_source = {tuple(item.source_object_ids): item for item in perimeter_items}
    for area in area_items:
        perimeter = perimeter_by_source.get(tuple(area.source_object_ids))
        rows.append([
            ",".join(area.source_object_ids),
            area.name.replace(" area", ""),
            area.quantity,
            area.unit,
            perimeter.quantity if perimeter else "",
            perimeter.unit if perimeter else "",
            area.confidence or "",
            "mock",
        ])
    return rows


def _opening_rows(quantities: list[QuantityItem]) -> list[list[object]]:
    rows = [["Type", "Count", "Unit", "Confidence", "Source Object IDs"]]
    for item in quantities:
        if item.category in {"door_count", "window_count"}:
            rows.append([item.name, item.quantity, item.unit, item.confidence or "", ",".join(item.source_object_ids)])
    return rows


def _wall_rows(quantities: list[QuantityItem]) -> list[list[object]]:
    rows = [["Object ID", "Name", "Length", "Unit", "Confidence", "Source"]]
    for item in quantities:
        if item.category in {"wall_length", "wall_length_total"}:
            rows.append([",".join(item.source_object_ids), item.name, item.quantity, item.unit, item.confidence or "", "mock"])
    return rows


def _object_rows(objects: list[DetectionObject]) -> list[list[object]]:
    rows = [["Object ID", "Type", "Source", "Geometry Type", "Label", "Confidence", "Geometry JSON"]]
    for item in objects:
        rows.append([item.id, item.type, item.source, item.geometry_type, item.label or "", item.confidence or "", json.dumps(to_plain(item.geometry))])
    return rows


def _content_types(sheet_count: int) -> str:
    overrides = "".join(f'<Override PartName="/xl/worksheets/sheet{i}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>' for i in range(1, sheet_count + 1))
    return '<?xml version="1.0" encoding="UTF-8"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>' + overrides + "</Types>"


def _rels() -> str:
    return '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>'


def _workbook_xml(names: list[str]) -> str:
    sheets = "".join(f'<sheet name="{name}" sheetId="{idx}" r:id="rId{idx}"/>' for idx, name in enumerate(names, start=1))
    return '<?xml version="1.0" encoding="UTF-8"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets>' + sheets + "</sheets></workbook>"


def _workbook_rels(sheet_count: int) -> str:
    rels = "".join(f'<Relationship Id="rId{i}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{i}.xml"/>' for i in range(1, sheet_count + 1))
    return '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">' + rels + "</Relationships>"

